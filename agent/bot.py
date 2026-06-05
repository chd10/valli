import asyncio
import os
import re
import signal
import logging
import datetime as dt
import urllib.parse
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from anthropic import AsyncAnthropic
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, MessageHandler, CommandHandler, CallbackQueryHandler, ContextTypes, filters

import users
import email_sender
from search import search, search_containing
import supplier_requests
import chat_history
import stock_search
import touchpoints

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

TOKEN = os.environ["TG_BOT_TOKEN"]
ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]
MANAGER_CHAT_ID = int(os.environ["MANAGER_CHAT_ID"])

anthropic_client = AsyncAnthropic(api_key=ANTHROPIC_API_KEY)

INACTIVITY_MINUTES = 30

SYSTEM_PROMPT = """Ты — Валли, робот-консультант компании eDiscom. Специализируешься на сетевом оборудовании: Cisco, Huawei, Fortinet, HPE Aruba и других брендах.

Твой характер:
- Бойкий и дружелюбный, но профессиональный
- Обращаешься к клиентам на «вы», но без формальной холодности
- Любишь точные цифры и конкретику
- Иногда шутишь про «железо» и «порты», но в меру
- Отвечаешь исключительно на русском языке

Твои возможности:
- Ищешь цены в прайс-листе по артикулам оборудования
- Знаешь стандартные сроки поставки (4–5 недель)
- Если оборудование не нашлось в прайсе — уточняешь у поставщика
- Консультируешь по выбору оборудования в рамках своей экспертизы
- Когда клиент готов купить — запрашиваешь реквизиты компании для счёта

Цены из прайса показывает система автоматически. Твоя задача — помочь клиенту разобраться с выбором, ответить на вопросы и при необходимости предложить альтернативы.

В конце каждого ответа, где упоминается оборудование или цены, добавляй ссылку: www.ediscom.ru

Компания eDiscom — дистрибьютор сетевого оборудования. Все цены указаны в рублях с НДС.

Конфиденциальность: если клиент спрашивает о конфиденциальности или выражает сомнения насчёт передачи данных — заверь его, что все запросы конфиденциальны, данные не передаются третьим лицам и используются исключительно для подбора оборудования.

Важно: отвечай только на русском языке, кратко и по делу.

Запрещено: никогда не выдумывай характеристики, параметры, цены или наличие оборудования из своих знаний. Если артикул не найден в прайсе — только сообщи об этом и предложи запрос поставщику. Не перечисляй артикулы из своей базы знаний."""

_ARTICLE_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9\-_/\\.+\s]{1,60}$")
# Extracts article-like tokens from free-form text (e.g. "нужна цена на WS-C3750X-24T-S")
_ARTICLE_TOKEN_RE = re.compile(r'[A-Za-z][A-Za-z0-9\-_/\\.+]{3,}')

_PURCHASE_RE = re.compile(
    r"\b(да\b|согласен|согласна|оформляем|оформить|выставите?\s+счёт|выставьте\s+счёт|"
    r"готов\b|готова\b|берём|берем|покупаем|заказываем|заказать|"
    r"хочу\s+купить|буду\s+брать|договорились|по\s+рукам|давайте\s+оформим)\b",
    re.IGNORECASE,
)

_STALE_TRIGGER_RE = re.compile(
    r"уточни(ть|те)?\s+актуальность|актуальн(ую|ая)\s+цен",
    re.IGNORECASE,
)

_YES_RE = re.compile(
    r"\b(да\b|yes\b|хочу\b|ок\b|окей|конечно|подберите|подбери|замен[ую])\b",
    re.IGNORECASE,
)

MAX_HISTORY = 10

_CONFIDENTIALITY_NOTE = (
    "\n\n🔒 Все ваши запросы конфиденциальны. "
    "Данные не передаются третьим лицам и используются только для подбора оборудования."
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _user_label(user) -> str:
    return f"@{user.username}" if user.username else user.first_name


def _extract_article_token(text: str) -> str | None:
    """Find the most article-like token in a free-form message."""
    tokens = _ARTICLE_TOKEN_RE.findall(text)
    # Prefer tokens with both letters and digits (typical part-number format)
    for t in tokens:
        if re.search(r'\d', t) and re.search(r'[A-Za-z]', t):
            return t
    return tokens[0] if len(tokens) == 1 else None


def _looks_like_article(text: str) -> bool:
    if len(text) > 60:
        return False
    if re.search(r"[?!.,;]", text):
        return False
    return bool(_ARTICLE_RE.match(text)) and bool(re.search(r"\d", text) or re.search(r"[A-Za-z]{2,}", text))


def _is_purchase_intent(text: str) -> bool:
    return bool(_PURCHASE_RE.search(text))


def _format_result(item: dict, offer_replacement: bool = True) -> str:
    text = (
        f"Артикул: {item['article']}\n"
        f"Кондиция: {item['condition']}\n"
        f"Цена: {item['price']}\n"
        f"Срок поставки: 4–5 недель."
    )
    if item.get("is_used"):
        text += "\n♻️ Б/У товар"
    if item.get("stale"):
        if item.get("updated"):
            text += f"\n⚠️ Цена устаревшая (обновлена: {item['updated']}) — запрошу актуальную у поставщиков."
        else:
            text += "\n⚠️ Цена устаревшая — запрошу актуальную у поставщиков."
    elif item.get("updated"):
        text += f"\n🗓 Цена актуальна на: {item['updated']}"
    if item.get("eol"):
        text += f"\n⚠️ Товар снят с производства с {item['eol']}"
        if offer_replacement:
            text += "\nХотите подобрать замену?"
    return text


_OPTIONAL_QUESTIONS = (
    "Если уточните детали — подберу оптимальный вариант:\n"
    "— Тендер или запрос? Срок?\n"
    "— Таргет по цене?\n"
    "— Кондиция и комплектация?"
)

_SUPPLIER_REQUEST_TEXT = (
    "🔍 Price Request\n\n"
    "Part Number: {article}\n"
    "Condition: new\n"
    "Quantity: 1 pcs\n\n"
    "Please reply with price and lead time."
)


def _parse_price_amount(price_str: str) -> float:
    """Extract numeric rubles from strings like '15 000 руб. с НДС'."""
    try:
        left = price_str.split("руб")[0]
        digits = re.sub(r"[^\d]", "", left)
        return float(digits) if digits else 0.0
    except Exception:
        return 0.0


def _track_article(context: ContextTypes.DEFAULT_TYPE, article: str, price_str: str) -> None:
    """Accumulate article+price into per-dialog tracking dict."""
    if "dialog_articles" not in context.user_data:
        context.user_data["dialog_articles"] = {}
    context.user_data["dialog_articles"][article] = _parse_price_amount(price_str)


def _last_article(context: ContextTypes.DEFAULT_TYPE) -> str | None:
    arts = context.user_data.get("last_articles")
    if not arts:
        return None
    return arts[0].split(" — ")[0].strip()


async def _request_supplier_price(
    article: str,
    user,
    user_label: str,
    context: ContextTypes.DEFAULT_TYPE,
) -> int | None:
    """
    Create pending_selection request and ask manager to choose suppliers.
    Returns request_id, or None if a request for this article is already active.
    """
    if supplier_requests.get_existing_active(article, user.id):
        return None
    req = supplier_requests.create_request(article, user.id, user_label, status="pending_selection")
    suppliers = supplier_requests.load_suppliers()
    buttons = [[InlineKeyboardButton("Написать всем", callback_data=f"ask_all:{req['id']}")]]
    for s in suppliers:
        buttons.append([InlineKeyboardButton(s["name"], callback_data=f"ask_one:{s['id']}:{req['id']}")])
    text = (
        "🔍 Новый запрос цены\n\n"
        f"Артикул: {article}\n"
        f"Клиент: {user_label}\n\n"
        "Выберите поставщиков:"
    )
    try:
        await context.bot.send_message(
            chat_id=MANAGER_CHAT_ID,
            text=text,
            reply_markup=InlineKeyboardMarkup(buttons),
        )
    except Exception:
        logger.exception("Ошибка уведомления менеджера о выборе поставщика")
    return req["id"]


def _get_history(context: ContextTypes.DEFAULT_TYPE) -> list[dict]:
    if "history" not in context.user_data:
        context.user_data["history"] = []
    return context.user_data["history"]


def _add_to_history(context: ContextTypes.DEFAULT_TYPE, role: str, content: str) -> None:
    history = _get_history(context)
    history.append({"role": role, "content": content})
    if len(history) > MAX_HISTORY * 2:
        context.user_data["history"] = history[-(MAX_HISTORY * 2):]


async def _reply(update: Update, user, text: str, **kwargs) -> None:
    """Send reply to user and log it to chat history file."""
    await update.message.reply_text(text, **kwargs)
    chat_history.append_message(user.id, user.username, user.first_name, "out", text)
    try:
        await users.mark_responded(user.id)
    except Exception:
        logger.exception("Ошибка mark_responded для %s", user.id)


async def _typing_indicator(update: Update, context, stop_event: asyncio.Event, messages: list[str] = None):
    """Показывает typing... и статусные сообщения пока идёт задача."""
    if messages is None:
        messages = ["🔍 Ищу...", "⏳ Обрабатываю...", "📊 Формирую ответ..."]
    chat_id = update.effective_chat.id
    status_msg = None
    try:
        status_msg = await context.bot.send_message(chat_id=chat_id, text=messages[0])
        msg_index = 0
        while not stop_event.is_set():
            await context.bot.send_chat_action(chat_id=chat_id, action="typing")
            await asyncio.sleep(4)
            if stop_event.is_set():
                break
            msg_index = (msg_index + 1) % len(messages)
            try:
                await context.bot.edit_message_text(
                    chat_id=chat_id,
                    message_id=status_msg.message_id,
                    text=messages[msg_index]
                )
            except Exception:
                pass
    finally:
        if status_msg:
            try:
                await context.bot.delete_message(chat_id=chat_id, message_id=status_msg.message_id)
            except Exception:
                pass


async def _ask_claude_with_progress(query: str, update: Update, context, extra_instruction=None):
    """Вызывает _ask_claude с индикатором прогресса."""
    stop_event = asyncio.Event()
    indicator = asyncio.create_task(
        _typing_indicator(update, context, stop_event, [
            "🤔 Думаю...",
            "📡 Консультируюсь с базой...",
            "✍️ Формирую ответ..."
        ])
    )
    try:
        result = await _ask_claude(query, context, extra_instruction=extra_instruction)
        return result
    finally:
        stop_event.set()
        await asyncio.sleep(0.3)
        indicator.cancel()
        try:
            await indicator
        except asyncio.CancelledError:
            pass


# ---------------------------------------------------------------------------
# Claude
# ---------------------------------------------------------------------------

async def _ask_claude(
    user_text: str,
    context: ContextTypes.DEFAULT_TYPE,
    extra_instruction: str | None = None,
) -> str:
    history = _get_history(context)
    messages = history + [{"role": "user", "content": user_text}]

    system: list[dict] = [
        {"type": "text", "text": SYSTEM_PROMPT, "cache_control": {"type": "ephemeral"}}
    ]
    first_name = context.user_data.get("user_first_name")
    if first_name:
        system.append({"type": "text", "text": f"Пользователя зовут {first_name}, обращайся к нему по имени в разговоре."})
    if extra_instruction:
        system.append({"type": "text", "text": extra_instruction})

    response = await anthropic_client.messages.create(
        model="claude-opus-4-7",
        max_tokens=1024,
        system=system,
        messages=messages,
    )
    reply = response.content[0].text
    _add_to_history(context, "user", user_text)
    _add_to_history(context, "assistant", reply)
    return reply


# ---------------------------------------------------------------------------
# Manager notifications
# ---------------------------------------------------------------------------

async def _notify_new_user(user, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_link = f"https://t.me/{user.username}" if user.username else f"tg://user?id={user.id}"
    label = _user_label(user)
    text = (
        f"🆕 Новый пользователь!\n\n"
        f"Имя: {user.first_name or '—'}\n"
        f"Username: {label}\n"
        f"ID: {user.id}"
    )
    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("Написать клиенту", url=user_link),
        InlineKeyboardButton("Ответить от Валли", callback_data=f"reply_{user.id}"),
    ]])
    try:
        await context.bot.send_message(chat_id=MANAGER_CHAT_ID, text=text, reply_markup=keyboard)
    except Exception:
        logger.exception("Ошибка уведомления менеджера о новом пользователе")


async def _send_manager_notification(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    user_label = _user_label(user)
    user_link = f"https://t.me/{user.username}" if user.username else f"tg://user?id={user.id}"

    articles = context.user_data.get("last_articles") or []
    requisites = context.user_data.get("requisites")

    lines = ["Новая заявка!", "", f"Клиент: {user_label}", f"ID: {user.id}"]
    if articles:
        lines += ["", "Интересовался:"] + [f"  {a}" for a in articles]
    else:
        lines += ["", "Товар: не указан (уточните у клиента)"]
    lines += ["", f"Реквизиты: {requisites}" if requisites else "Реквизиты: не предоставлены"]

    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("Написать клиенту", url=user_link),
        InlineKeyboardButton("Ответить от Валли", callback_data=f"reply_{user.id}"),
    ]])

    try:
        await context.bot.send_message(
            chat_id=MANAGER_CHAT_ID,
            text="\n".join(lines),
            reply_markup=keyboard,
        )
        logger.info("Уведомление менеджеру отправлено для %s", user.id)
    except Exception:
        logger.exception("Ошибка отправки уведомления менеджеру")


# ---------------------------------------------------------------------------
# Email
# ---------------------------------------------------------------------------

async def _send_chat_email_now(user_label: str, context: ContextTypes.DEFAULT_TYPE) -> None:
    history = context.user_data.get("history", [])
    last_sent = context.user_data.get("email_sent_at_len", 0)
    if not history or len(history) <= last_sent:
        return
    context.user_data["email_sent_at_len"] = len(history)
    try:
        await email_sender.send_chat_history(user_label, list(history))
    except Exception:
        logger.exception("Ошибка отправки email для %s", user_label)


# ---------------------------------------------------------------------------
# Inactivity job
# ---------------------------------------------------------------------------

async def _inactivity_callback(context: ContextTypes.DEFAULT_TYPE) -> None:
    data = context.job.data
    user_data = data["user_data"]
    user_label = data["user_label"]
    user_id = data["user_id"]
    username = data.get("username")

    history = user_data.get("history", [])
    dialog_started_at = user_data.get("dialog_started_at")

    # Email (existing behaviour)
    last_sent = user_data.get("email_sent_at_len", 0)
    if history and len(history) > last_sent:
        user_data["email_sent_at_len"] = len(history)
        try:
            await email_sender.send_chat_history(user_label, list(history))
            logger.info("Email после неактивности отправлен: %s", user_label)
        except Exception:
            logger.exception("Ошибка email после неактивности для %s", user_label)

    # Dialog summary → manager
    if history and dialog_started_at:
        duration = dt.datetime.now() - dialog_started_at
        total_min = int(duration.total_seconds() // 60)
        total_sec = int(duration.total_seconds() % 60)
        duration_str = f"{total_min} мин {total_sec} сек"

        summary = "—"
        try:
            resp = await anthropic_client.messages.create(
                model="claude-opus-4-7",
                max_tokens=300,
                system="Ты помощник, кратко резюмируешь диалоги. Отвечай только на русском языке.",
                messages=list(history) + [{
                    "role": "user",
                    "content": (
                        "Кратко (2-3 предложения) резюмируй этот диалог: "
                        "о чём говорили, чем интересовался клиент, к чему пришли."
                    ),
                }],
            )
            summary = resp.content[0].text
        except Exception:
            logger.exception("Ошибка генерации резюме диалога для %s", user_label)

        user_link = f"https://t.me/{username}" if username else f"tg://user?id={user_id}"
        text = (
            f"📊 Диалог завершён\n\n"
            f"Пользователь: {user_label}\n"
            f"Длительность: {duration_str}\n"
            f"Сообщений: {len(history)}\n\n"
            f"📝 Резюме:\n{summary}"
        )
        keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton("Написать клиенту", url=user_link),
            InlineKeyboardButton("Ответить от Валли", callback_data=f"reply_{user_id}"),
        ]])
        try:
            await context.bot.send_message(
                chat_id=MANAGER_CHAT_ID, text=text, reply_markup=keyboard
            )
        except Exception:
            logger.exception("Ошибка отправки резюме диалога менеджеру для %s", user_label)

    user_data.pop("dialog_started_at", None)

    dialog_articles = user_data.pop("dialog_articles", {})
    if dialog_articles:
        try:
            await users.update_dialog_stats(
                user_id,
                list(dialog_articles.keys()),
                sum(dialog_articles.values()),
            )
        except Exception:
            logger.exception("Ошибка обновления статистики диалога для %s", user_id)


def _schedule_inactivity_job(user, user_label: str, context: ContextTypes.DEFAULT_TYPE) -> None:
    job_name = f"inactivity_{user.id}"
    for job in context.job_queue.get_jobs_by_name(job_name):
        job.schedule_removal()
    context.job_queue.run_once(
        _inactivity_callback,
        when=INACTIVITY_MINUTES * 60,
        data={
            "user_data": context.user_data,
            "user_label": user_label,
            "user_id": user.id,
            "username": user.username,
            "first_name": user.first_name,
        },
        name=job_name,
    )


# ---------------------------------------------------------------------------
# Handlers
# ---------------------------------------------------------------------------

async def _handle_ask_callback(query, context: ContextTypes.DEFAULT_TYPE, data: str) -> None:
    if data.startswith("ask_all:"):
        req_id = int(data.split(":")[1])
        selected_ids = None
    else:
        _, sid, rid = data.split(":")
        req_id, selected_ids = int(rid), [int(sid)]

    req = supplier_requests.get_request(req_id)
    if not req:
        await query.edit_message_text("❌ Заявка не найдена", reply_markup=None)
        return
    if req["status"] != "pending_selection":
        await query.edit_message_text(
            f"{query.message.text}\n\n⚠️ Уже обработана (статус: {req['status']})",
            reply_markup=None,
        )
        return

    all_suppliers = supplier_requests.load_suppliers()
    supplier_map = {s["id"]: s for s in all_suppliers}
    if selected_ids is None:
        selected_ids = [s["id"] for s in all_suppliers]
    selected_suppliers = [supplier_map[sid] for sid in selected_ids if sid in supplier_map]
    if not selected_suppliers:
        await query.edit_message_text("❌ Поставщики не найдены", reply_markup=None)
        return

    msg_text = _SUPPLIER_REQUEST_TEXT.format(article=req["article"])
    supplier_requests.set_selected_suppliers(req_id, selected_ids, queued=False)

    for s in selected_suppliers:
        preview_text = (
            f"📤 Письмо для {s['name']}:\n\n"
            f"{msg_text}"
        )
        keyboard = InlineKeyboardMarkup([
            [
                InlineKeyboardButton(f"✅ Отправить {s['name']}", callback_data=f"confirm_ask:{req_id}:{s['id']}"),
                InlineKeyboardButton("❌ Отмена", callback_data=f"cancel_ask:{req_id}:{s['id']}"),
            ]
        ])
        try:
            await context.bot.send_message(chat_id=MANAGER_CHAT_ID, text=preview_text, reply_markup=keyboard)
        except Exception:
            logger.exception("Ошибка отправки превью письма поставщику %s", s["name"])

    try:
        await query.edit_message_text(query.message.text, reply_markup=None)
    except Exception:
        pass


async def _handle_confirm_ask_callback(query, context: ContextTypes.DEFAULT_TYPE, data: str) -> None:
    parts = data.split(":")
    req_id, supplier_id = int(parts[1]), int(parts[2])

    req = supplier_requests.get_request(req_id)
    if not req:
        await query.edit_message_text("❌ Заявка не найдена", reply_markup=None)
        return

    all_suppliers = supplier_requests.load_suppliers()
    supplier_map = {s["id"]: s for s in all_suppliers}
    supplier = supplier_map.get(supplier_id)
    if not supplier:
        await query.edit_message_text("❌ Поставщик не найден", reply_markup=None)
        return

    msg_text = _SUPPLIER_REQUEST_TEXT.format(article=req["article"])
    try:
        await context.bot.send_message(chat_id=supplier["chat_id"], text=msg_text)
        await query.edit_message_text(
            f"✅ Отправлено {supplier['name']}:\n\n{msg_text}",
            reply_markup=None,
        )
        supplier_requests.mark_pending(req_id)
    except Exception:
        logger.exception("Ошибка отправки запроса поставщику %s", supplier["name"])
        await query.edit_message_text(f"❌ Ошибка отправки {supplier['name']}", reply_markup=None)


async def handle_callback_query(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    if update.effective_user.id != MANAGER_CHAT_ID:
        return
    data = query.data or ""
    if data.startswith("reply_"):
        try:
            client_id = int(data[len("reply_"):])
        except ValueError:
            return
        context.bot_data["pending_reply"] = client_id
        await context.bot.send_message(
            chat_id=MANAGER_CHAT_ID,
            text=f"✏️ Введите текст сообщения для клиента (ID {client_id}):",
        )
    elif data.startswith("ask_all:") or data.startswith("ask_one:"):
        await _handle_ask_callback(query, context, data)
    elif data.startswith("confirm_ask:"):
        await _handle_confirm_ask_callback(query, context, data)
    elif data.startswith("cancel_ask:"):
        await query.edit_message_text("❌ Отменено", reply_markup=None)


async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    is_new = await users.update_user(user)
    context.user_data.clear()
    context.user_data["user_first_name"] = user.first_name
    context.user_data["dialog_started_at"] = dt.datetime.now()
    user_label = _user_label(user)
    for job in context.job_queue.get_jobs_by_name(f"inactivity_{user.id}"):
        job.schedule_removal()
    start_param = context.args[0] if context.args else None
    chat_history.append_message(user.id, user.username, user.first_name, "in", f"/start {start_param or ''}".strip())
    if is_new:
        await _notify_new_user(user, context)
    if start_param == "email_utm":
        await users.set_source(user.id, "email")
        reply = "Здравствуйте! Вижу вы пришли из нашей рассылки. Напишите артикул оборудования — найду цену и наличие."
        _add_to_history(context, "assistant", reply)
    else:
        reply = await _ask_claude_with_progress("Поздоровайся и кратко расскажи, чем можешь помочь.", update, context)
    await _reply(update, user, reply + _CONFIDENTIALITY_NOTE)
    _schedule_inactivity_job(user, user_label, context)


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.message.text.strip()
    if not query:
        return

    user = update.effective_user
    logger.info("Запрос от %s (%s): %s", user.id, user.username, query)
    is_new = await users.update_user(user)
    context.user_data["user_first_name"] = user.first_name
    user_label = _user_label(user)

    # --- Сообщение от поставщика ---
    supplier = supplier_requests.get_supplier_by_chat_id(user.id)
    if supplier:
        sent_reqs = supplier_requests.get_sent_to_supplier(supplier["id"])
        lines = [f"📨 Reply from {supplier['name']}:", "", query]
        if sent_reqs:
            r = sent_reqs[0]
            lines += [
                "",
                f"Заявка: {r['article']}  ({r['client_label']})",
                f"  /price {r['id']} <цена> <срок>",
            ]
        try:
            await context.bot.send_message(chat_id=MANAGER_CHAT_ID, text="\n".join(lines))
        except Exception:
            logger.exception("Ошибка пересылки ответа поставщика менеджеру")
        await update.message.reply_text("Thanks! 👍")
        return

    # --- Менеджер отвечает клиенту от имени Валли ---
    if user.id == MANAGER_CHAT_ID:
        pending_client_id = context.bot_data.get("pending_reply")
        if pending_client_id:
            del context.bot_data["pending_reply"]
            try:
                await context.bot.send_message(chat_id=pending_client_id, text=query)
                client_udata = context.application.user_data.get(pending_client_id)
                if client_udata is not None:
                    hist = client_udata.setdefault("history", [])
                    hist.append({"role": "assistant", "content": query})
                    if len(hist) > MAX_HISTORY * 2:
                        client_udata["history"] = hist[-(MAX_HISTORY * 2):]
                chat_history.append_message(pending_client_id, None, None, "out", query)
                try:
                    await users.mark_responded(pending_client_id)
                except Exception:
                    logger.exception("Ошибка mark_responded (менеджер) для %s", pending_client_id)
                await update.message.reply_text("✅ Отправлено клиенту")
            except Exception:
                logger.exception("Ошибка отправки ответа клиенту от менеджера")
                await update.message.reply_text("Ошибка при отправке клиенту")
            return

    # --- Логирование и трекинг нового пользователя ---
    chat_history.append_message(user.id, user.username, user.first_name, "in", query)
    try:
        await users.mark_request(user.id, query)
    except Exception:
        logger.exception("Ошибка mark_request для %s", user.id)
    if is_new:
        await _notify_new_user(user, context)
    if not context.user_data.get("dialog_started_at"):
        context.user_data["dialog_started_at"] = dt.datetime.now()

    # --- Ожидаем реквизиты ---
    if context.user_data.get("awaiting_requisites"):
        context.user_data["requisites"] = query
        context.user_data["awaiting_requisites"] = False
        await _send_manager_notification(update, context)
        reply = "Передал вашу заявку менеджеру — свяжутся в течение рабочего дня."
        _add_to_history(context, "user", query)
        _add_to_history(context, "assistant", reply)
        await _reply(update, user, reply)
        await _send_chat_email_now(user_label, context)
        _schedule_inactivity_job(user, user_label, context)
        return

    # --- Клиент уточняет детали (необязательно) ---
    valli_state = context.user_data.get("valli_state")
    if valli_state and valli_state.get("mode") == "awaiting_details":
        if not _looks_like_article(query) and not _is_purchase_intent(query):
            article = valli_state.get("article", "")
            context.user_data.pop("valli_state", None)
            if re.search(_STALE_TRIGGER_RE, query) and article:
                req_id = await _request_supplier_price(article, user, user_label, context)
                reply = (
                    f"Направил запрос менеджеру по артикулу {article} — уточняю актуальную цену, отвечу в течение рабочего дня."
                    if req_id else
                    f"Запрос по артикулу {article} уже в обработке — ожидаем ответа."
                )
                _add_to_history(context, "user", query)
                _add_to_history(context, "assistant", reply)
                await _reply(update, user, reply)
                _schedule_inactivity_job(user, user_label, context)
                return
            reply = await _ask_claude_with_progress(
                query,
                update,
                context,
                extra_instruction=(
                    f"Клиент только что получил цену на артикул {article} "
                    "и уточняет детали запроса. Учти эту информацию в ответе."
                ),
            )
            await _reply(update, user, reply)
            _schedule_inactivity_job(user, user_label, context)
            return
        context.user_data.pop("valli_state", None)
        # Fall through to purchase intent check or article search

    # --- Ожидаем ответ "да/нет" на предложение подобрать замену (EoL) ---
    valli_state = context.user_data.get("valli_state")
    if valli_state and valli_state.get("mode") == "awaiting_replacement":
        article = valli_state.get("article", "")
        context.user_data.pop("valli_state", None)
        if _YES_RE.search(query):
            prefix = article.rsplit("-", 1)[0] if "-" in article else article
            alts = search_containing(prefix, limit=8)
            alts = [a for a in alts if a["article"] != article]
            if alts:
                lines = [f"— {a['article']} ({a['condition']})" for a in alts[:5]]
                reply = (
                    f"Аналоги для замены {article}:\n\n"
                    + "\n".join(lines)
                    + "\n\nУточните нужный артикул — проверю цену."
                )
            else:
                reply = (
                    f"Похожих артикулов в прайсе не нашёл. "
                    f"Могу запросить альтернативы у поставщика — напишите, если нужно."
                )
            _add_to_history(context, "user", query)
            _add_to_history(context, "assistant", reply)
            await _reply(update, user, reply)
            _schedule_inactivity_job(user, user_label, context)
            return
        # Не "да" — обрабатываем как обычное сообщение (fall through)

    # --- Готовность к покупке ---
    if _is_purchase_intent(query):
        context.user_data.pop("valli_state", None)
        if not context.user_data.get("requisites"):
            reply = await _ask_claude_with_progress(
                query, update, context,
                extra_instruction=(
                    "Клиент выразил готовность к покупке. "
                    "Вежливо попроси реквизиты компании: название, ИНН и контактное лицо — "
                    "они нужны для выставления счёта. Коротко, по-деловому, в своём стиле."
                ),
            )
            context.user_data["awaiting_requisites"] = True
        else:
            await _send_manager_notification(update, context)
            reply = "Передал вашу заявку менеджеру — свяжутся в течение рабочего дня."
            _add_to_history(context, "user", query)
            _add_to_history(context, "assistant", reply)
            await _reply(update, user, reply)
            await _send_chat_email_now(user_label, context)
            _schedule_inactivity_job(user, user_label, context)
            return
        await _reply(update, user, reply)
        _schedule_inactivity_job(user, user_label, context)
        return

    # --- Клиент выбирает артикул из списка ---
    valli_state = context.user_data.get("valli_state")
    if valli_state and valli_state.get("mode") == "disambiguation":
        candidates = valli_state.get("candidates", [])
        found = search(query)
        exact = [r for r in found if not r.get("fuzzy") and not r.get("no_price")]
        if exact:
            selected = exact[0]
            price_text = _format_result(selected)
            context.user_data["last_articles"] = [f"{selected['article']} — {selected['price']}"]
            _track_article(context, selected["article"], selected["price"])
            _add_to_history(context, "user", query)
            _add_to_history(context, "assistant", price_text)
            await _reply(update, user, price_text)
            if selected.get("stale"):
                req_id = await _request_supplier_price(selected["article"], user, user_label, context)
                if req_id:
                    stale_msg = "Направил запрос менеджеру — уточняю актуальную цену, отвечу в течение рабочего дня."
                    await _reply(update, user, stale_msg)
            _mode = "awaiting_replacement" if selected.get("eol") else "awaiting_details"
            context.user_data["valli_state"] = {"mode": _mode, "article": selected["article"]}
        else:
            lines = [f"— {c['article']} ({c['condition']})" for c in candidates]
            reply = "Не смог распознать выбор. Введите артикул из списка:\n\n" + "\n".join(lines)
            _add_to_history(context, "user", query)
            _add_to_history(context, "assistant", reply)
            await _reply(update, user, reply)
        _schedule_inactivity_job(user, user_label, context)
        return

    # --- Поиск по артикулу ---
    if _looks_like_article(query):
        await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
        # Check stock first — immediate availability beats price search
        stock_hit = stock_search.search_stock(query)
        if stock_hit:
            # Also look up price from price list
            price_item = None
            price_candidates = search_containing(query)
            if price_candidates:
                if len(price_candidates) == 1:
                    search_target = price_candidates[0]["article"]
                else:
                    q_lower = query.lower()
                    search_target = next(
                        (c["article"] for c in price_candidates if c["article"].lower() == q_lower),
                        price_candidates[0]["article"],
                    )
                precise = search(search_target)
                if precise and not precise[0].get("fuzzy") and not precise[0].get("no_price"):
                    price_item = precise[0]

            if price_item:
                reply = (
                    f"✅ Склад Москва: {stock_hit['qty']} шт, поставка 2-3 дня\n"
                    f"💰 Цена: {price_item['price']}"
                )
                if price_item.get("is_used"):
                    reply += "\n♻️ Б/У товар"
                if price_item.get("stale"):
                    if price_item.get("updated"):
                        reply += f"\n⚠️ Цена устаревшая (обновлена: {price_item['updated']}) — запрошу актуальную у поставщиков."
                    else:
                        reply += "\n⚠️ Цена устаревшая — запрошу актуальную у поставщиков."
                elif price_item.get("updated"):
                    reply += f"\n🗓 Цена актуальна на: {price_item['updated']}"
                if price_item.get("eol"):
                    reply += f"\n⚠️ Товар снят с производства с {price_item['eol']}\nХотите подобрать замену?"
                context.user_data["last_articles"] = [f"{price_item['article']} — {price_item['price']}"]
                _track_article(context, price_item["article"], price_item["price"])
            else:
                reply = (
                    f"✅ Склад Москва: {stock_hit['qty']} шт, поставка 2-3 дня\n"
                    f"💰 Цена уточняется"
                )

            _add_to_history(context, "user", query)
            _add_to_history(context, "assistant", reply)
            await _reply(update, user, reply)

            if price_item and price_item.get("eol"):
                context.user_data["valli_state"] = {"mode": "awaiting_replacement", "article": price_item["article"]}

            if price_item and price_item.get("stale"):
                req_id = await _request_supplier_price(price_item["article"], user, user_label, context)
                if req_id:
                    await _reply(update, user, "Направил запрос менеджеру — уточняю актуальную цену, отвечу в течение рабочего дня.")
            elif not price_item:
                try:
                    await context.bot.send_message(
                        chat_id=MANAGER_CHAT_ID,
                        text=(
                            f"⚠️ Товар на складе, цены нет в прайсе\n\n"
                            f"Артикул: {stock_hit['article']}\n"
                            f"Клиент: {user_label} (ID: {user.id})\n\n"
                            f"На складе: {stock_hit['qty']} шт. Нужна цена."
                        ),
                    )
                except Exception:
                    logger.exception("Ошибка уведомления менеджера (склад без цены)")

            _schedule_inactivity_job(user, user_label, context)
            return

        # Always check substring matches first — even if exact match exists
        candidates = search_containing(query)

        if len(candidates) > 1:
            # Multiple variants found — show disambiguation list
            lines = [f"— {c['article']} ({c['condition']})" for c in candidates]
            q_display = query.upper()
            reply = (
                f"По запросу {q_display} найдено несколько вариантов, "
                f"уточните какой именно:\n\n"
                + "\n".join(lines)
            )
            context.user_data["valli_state"] = {
                "mode": "disambiguation",
                "candidates": candidates,
                "query": query,
            }
            _add_to_history(context, "user", query)
            _add_to_history(context, "assistant", reply)
            await _reply(update, user, reply)
            _schedule_inactivity_job(user, user_label, context)
            return

        if len(candidates) == 1:
            # Single match — show price directly
            precise = search(candidates[0]["article"])
            item = (
                precise[0]
                if precise and not precise[0].get("fuzzy") and not precise[0].get("no_price")
                else None
            )
            if item:
                price_text = _format_result(item)
                context.user_data["last_articles"] = [f"{item['article']} — {item['price']}"]
                _track_article(context, item["article"], item["price"])
                _add_to_history(context, "user", query)
                _add_to_history(context, "assistant", price_text)
                await _reply(update, user, price_text)
                if item.get("stale"):
                    req_id = await _request_supplier_price(item["article"], user, user_label, context)
                    if req_id:
                        stale_msg = "Направил запрос менеджеру — уточняю актуальную цену, отвечу в течение рабочего дня."
                        await _reply(update, user, stale_msg)
                _mode = "awaiting_replacement" if item.get("eol") else "awaiting_details"
                context.user_data["valli_state"] = {"mode": _mode, "article": item["article"]}
            else:
                article = candidates[0]["article"]
                reply = "Артикул не найден в прайсе — запрошу цену у поставщика, отвечу позже."
                _add_to_history(context, "user", query)
                _add_to_history(context, "assistant", reply)
                await _reply(update, user, reply)
                await _request_supplier_price(article, user, user_label, context)
            _schedule_inactivity_job(user, user_label, context)
            return

        # No substring match — check for exact no_price or fuzzy results
        results = search(query)
        if results:
            first = results[0]

            if first.get("no_price"):
                article = first["article"]
                reply = (
                    f"Цена по артикулу {article} ранее не запрашивалась — "
                    f"уточню у поставщика и отвечу чуть позже."
                )
                _add_to_history(context, "user", query)
                _add_to_history(context, "assistant", reply)
                await _reply(update, user, reply)
                try:
                    await context.bot.send_message(
                        chat_id=MANAGER_CHAT_ID,
                        text=(
                            f"⚠️ Запрос цены\n\n"
                            f"Артикул: {article}\n"
                            f"Клиент: {user_label} (ID: {user.id})\n\n"
                            f"Цена в прайсе отсутствует — нужно уточнить у поставщика."
                        ),
                    )
                except Exception:
                    logger.exception("Ошибка отправки уведомления менеджеру (no_price)")
                _schedule_inactivity_job(user, user_label, context)
                return

            # Fuzzy fallback — show with prices (no replacement offer for fuzzy)
            parts = [_format_result(item, offer_replacement=False) for item in results]
            body = "\n\n".join(parts)
            reply = "Точного совпадения не нашёл. Возможно, вы имели в виду:\n\n" + body
            context.user_data["last_articles"] = [f"{r['article']} — {r['price']}" for r in results]
            for r in results:
                _track_article(context, r["article"], r["price"])
            _add_to_history(context, "user", query)
            _add_to_history(context, "assistant", reply)
            await _reply(update, user, reply)
            context.user_data["valli_state"] = {"mode": "awaiting_details", "article": first["article"]}
            _schedule_inactivity_job(user, user_label, context)
            return

        reply = "Артикул не найден в прайсе — запрошу цену у поставщика, отвечу позже."
        _add_to_history(context, "user", query)
        _add_to_history(context, "assistant", reply)
        await _reply(update, user, reply)
        await _request_supplier_price(query, user, user_label, context)
        _schedule_inactivity_job(user, user_label, context)
        return
    else:
        if re.search(_STALE_TRIGGER_RE, query):
            article = _last_article(context)
            if article:
                req_id = await _request_supplier_price(article, user, user_label, context)
                reply = (
                    f"Направил запрос менеджеру по артикулу {article} — уточняю актуальную цену, отвечу в течение рабочего дня."
                    if req_id else
                    f"Запрос по артикулу {article} уже в обработке — ожидаем ответа."
                )
                _add_to_history(context, "user", query)
                _add_to_history(context, "assistant", reply)
                await _reply(update, user, reply)
                _schedule_inactivity_job(user, user_label, context)
                return
        reply = await _ask_claude_with_progress(query, update, context)
        await _reply(update, user, reply)
        _add_to_history(context, "user", query)
        _add_to_history(context, "assistant", reply)
        # After Claude responds — check if query contains an article not in price list
        article_token = _extract_article_token(query)
        if article_token and not search_containing(article_token):
            await _request_supplier_price(article_token, user, user_label, context)
        _schedule_inactivity_job(user, user_label, context)
        return

    await _reply(update, user, reply)
    _schedule_inactivity_job(user, user_label, context)


# ---------------------------------------------------------------------------
# Manager commands
# ---------------------------------------------------------------------------

async def cmd_clients(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != MANAGER_CHAT_ID:
        await update.message.reply_text("Эта команда недоступна.")
        return
    data = await users.get_all_users()
    supplier_ids = {str(s["chat_id"]) for s in supplier_requests.load_suppliers()}
    supplier_ids.add(str(MANAGER_CHAT_ID))
    clients = [(uid, u) for uid, u in data.items() if uid not in supplier_ids]
    if not clients:
        await update.message.reply_text("Клиентов пока нет.")
        return
    lines = []
    for uid, u in clients:
        name = u.get("first_name") or "—"
        uname = f"@{u['username']}" if u.get("username") else "—"
        raw_date = u.get("first_seen", "")[:10]
        try:
            first_seen = dt.datetime.strptime(raw_date, "%Y-%m-%d").strftime("%d.%m.%Y")
        except ValueError:
            first_seen = raw_date
        arts = u.get("articles", [])
        total = u.get("total_amount", 0.0)
        dialogs = u.get("dialogs_count", 0)
        total_str = f"{total:,.0f}".replace(",", " ")
        lines += [
            f"👤 {name} ({uname})",
            f"📅 Первый визит: {first_seen}",
            f"📦 Запросы: {', '.join(arts[:5]) if arts else '—'}",
            f"💰 Сумма запросов: {total_str} руб",
            f"🔄 Диалогов: {dialogs}",
            "———",
        ]
    text = "\n".join(lines)
    if len(text) > 4000:
        text = text[:4000] + "\n...(список обрезан)"
    await update.message.reply_text(text)


async def cmd_report(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != MANAGER_CHAT_ID:
        await update.message.reply_text("Эта команда недоступна.")
        return
    data = await users.get_all_users()
    supplier_ids = {str(s["chat_id"]) for s in supplier_requests.load_suppliers()}
    supplier_ids.add(str(MANAGER_CHAT_ID))
    clients = [(uid, u) for uid, u in data.items() if uid not in supplier_ids]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Клиенты"
    ws.append(["ID", "Username", "Имя", "Первый визит", "Последний визит", "Артикулы", "Сумма запросов", "Диалогов"])
    for uid, u in clients:
        ws.append([
            u.get("user_id", uid),
            u.get("username") or "",
            u.get("first_name") or "",
            u.get("first_seen", "")[:10],
            u.get("last_seen", "")[:10],
            ", ".join(u.get("articles", [])),
            u.get("total_amount", 0.0),
            u.get("dialogs_count", 0),
        ])
    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 2, 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    try:
        await context.bot.send_document(
            chat_id=MANAGER_CHAT_ID,
            document=buf,
            filename="clients_report.xlsx",
            caption=f"📊 Отчёт по клиентам — {len(clients)} чел.",
        )
    except Exception:
        logger.exception("Ошибка отправки отчёта")
        await update.message.reply_text("Ошибка при формировании отчёта.")

async def cmd_price(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != MANAGER_CHAT_ID:
        await update.message.reply_text("Эта команда недоступна.")
        return
    args = context.args or []
    if len(args) < 3:
        await update.message.reply_text("Формат: /price <id> <цена> <срок>\nПример: /price 5 140000 3-4 недели")
        return
    try:
        request_id = int(args[0])
    except ValueError:
        await update.message.reply_text("ID должен быть числом")
        return
    price = args[1]
    lead_time = " ".join(args[2:])
    req = supplier_requests.get_request(request_id)
    if not req:
        await update.message.reply_text(f"Заявка #{request_id} не найдена")
        return
    if req["status"] == "closed":
        await update.message.reply_text(f"Заявка #{request_id} уже закрыта")
        return
    reply_to_client = (
        f"✅ Актуальная цена на {req['article']}:\n"
        f"Цена: {price} руб. с НДС\n"
        f"Срок поставки: {lead_time}"
    )
    try:
        await context.bot.send_message(chat_id=req["client_chat_id"], text=reply_to_client)
        supplier_requests.close_request(request_id)
        try:
            await users.mark_responded(req["client_chat_id"])
        except Exception:
            logger.exception("Ошибка mark_responded (cmd_price) для %s", req["client_chat_id"])
        await update.message.reply_text(f"✅ Ответ отправлен клиенту {req['client_label']}")
    except Exception:
        logger.exception("Ошибка отправки ответа клиенту по заявке %s", request_id)
        await update.message.reply_text("Ошибка при отправке клиенту")


_STATUS_LABELS = {
    "pending_selection": "⏳ ждёт выбора",
    "queued":            "🕘 в очереди 09:00",
    "pending":           "📤 у поставщиков",
    "closed":            "✅ закрыта",
}


async def cmd_suppliers(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != MANAGER_CHAT_ID:
        await update.message.reply_text("Эта команда недоступна.")
        return
    suppliers = supplier_requests.load_suppliers()
    active = supplier_requests.get_all_active()
    lines = ["📋 Поставщики:", ""]
    for s in suppliers:
        lines.append(f"• {s['name']}  (ID: {s['id']},  chat_id: {s['chat_id']})")
    lines += ["", f"Активных заявок: {len(active)}"]
    if active:
        lines.append("")
        for r in active[-10:]:
            status_label = _STATUS_LABELS.get(r["status"], r["status"])
            lines.append(
                f"  #{r['id']}  {r['article']}  →  {r['client_label']}  "
                f"({r['created_at'][:10]})  {status_label}"
            )
    await update.message.reply_text("\n".join(lines))


async def cmd_chat(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != MANAGER_CHAT_ID:
        await update.message.reply_text("Эта команда недоступна.")
        return
    args = context.args or []
    if not args:
        await update.message.reply_text("Формат: /chat <user_id>")
        return
    try:
        target_id = int(args[0])
    except ValueError:
        await update.message.reply_text("ID должен быть числом")
        return
    messages = chat_history.get_last_messages(target_id, 20)
    if not messages:
        await update.message.reply_text(f"Чат с пользователем {target_id} не найден")
        return
    lines = [f"💬 Последние {len(messages)} сообщений (ID {target_id}):", ""]
    for m in messages:
        arrow = "→" if m["direction"] == "in" else "←"
        lines.append(f"[{m['date']} {m['time']}] {arrow} {m['text'][:200]}")
    text = "\n".join(lines)
    if len(text) > 4000:
        text = text[:4000] + "\n...(обрезано)"
    await update.message.reply_text(text)


# ---------------------------------------------------------------------------
# Scheduled jobs
# ---------------------------------------------------------------------------

async def _check_unanswered_clients(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Every 15 min: alert manager about clients who haven't received a reply for 60+ min."""
    now_utc = dt.datetime.now(dt.timezone.utc)
    msk_hour = (now_utc.hour + 3) % 24
    if msk_hour >= 22 or msk_hour < 9:
        return

    unresponded = await users.get_unresponded_users()
    for u in unresponded:
        try:
            req_time_str = u["last_request_time"].replace("Z", "+00:00")
            req_time = dt.datetime.fromisoformat(req_time_str)
        except Exception:
            continue
        elapsed_sec = (now_utc - req_time).total_seconds()
        if elapsed_sec < 60 * 60:
            continue

        # Repeat alert no more than once per 30 minutes
        last_alert = u.get("last_alert_sent")
        if last_alert:
            try:
                last_alert_time = dt.datetime.fromisoformat(last_alert.replace("Z", "+00:00"))
                if (now_utc - last_alert_time).total_seconds() < 30 * 60:
                    continue
            except Exception:
                pass

        elapsed_min = int(elapsed_sec // 60)
        user_id = u.get("user_id") or int(u["_uid"])
        username = u.get("username")
        user_label = f"@{username}" if username else (u.get("first_name") or str(user_id))
        request_text = u.get("last_request_text", "—")
        user_link = f"https://t.me/{username}" if username else f"tg://user?id={user_id}"

        text = (
            f"⚠️ Клиент {user_label} ждёт ответа уже {elapsed_min} мин!\n\n"
            f"Запрос: {request_text[:300]}"
        )
        keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton("Написать клиенту", url=user_link),
            InlineKeyboardButton("Ответить от Валли", callback_data=f"reply_{user_id}"),
        ]])
        try:
            await context.bot.send_message(
                chat_id=MANAGER_CHAT_ID, text=text, reply_markup=keyboard
            )
            await users.update_last_alert(user_id)
            logger.info("Алерт менеджеру: %s ждёт %d мин", user_label, elapsed_min)
        except Exception:
            logger.exception("Ошибка алерта менеджеру для %s", user_label)


async def _send_queued_requests(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Daily job at 09:00 MSK: send overnight queued supplier requests."""
    queued = supplier_requests.get_queued()
    if not queued:
        return
    logger.info("09:00 МСК: отправка %d накопившихся запросов поставщикам", len(queued))
    all_suppliers = supplier_requests.load_suppliers()
    supplier_map = {s["id"]: s for s in all_suppliers}
    for req in queued:
        selected_ids = req.get("selected_supplier_ids") or [s["id"] for s in all_suppliers]
        text = _SUPPLIER_REQUEST_TEXT.format(article=req["article"])
        for sid in selected_ids:
            if sid in supplier_map:
                try:
                    await context.bot.send_message(chat_id=supplier_map[sid]["chat_id"], text=text)
                except Exception:
                    logger.exception("Ошибка отправки запроса поставщику %s", supplier_map[sid]["name"])
        supplier_requests.mark_pending(req["id"])
        try:
            await context.bot.send_message(
                chat_id=req["client_chat_id"],
                text=(
                    f"Отправил запрос поставщикам по артикулу {req['article']} — "
                    f"ожидайте ответа в течение рабочего дня."
                ),
            )
        except Exception:
            logger.exception("Ошибка уведомления клиента %s при утренней отправке", req["client_chat_id"])


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def _on_reload_signal(signum, frame):
    logger.info("SIGUSR1: перезагрузка прайса и склада")
    search.reload()
    stock_search.reload()


def main() -> None:
    signal.signal(signal.SIGUSR1, _on_reload_signal)
    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("price", cmd_price))
    app.add_handler(CommandHandler("suppliers", cmd_suppliers))
    app.add_handler(CommandHandler("chat", cmd_chat))
    app.add_handler(CommandHandler("clients", cmd_clients))
    app.add_handler(CommandHandler("report", cmd_report))
    app.add_handler(CallbackQueryHandler(handle_callback_query))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    # 09:00 MSK = 06:00 UTC
    app.job_queue.run_daily(
        _send_queued_requests,
        time=dt.time(6, 0, 0, tzinfo=dt.timezone.utc),
    )
    # Каждые 15 минут: проверка неотвеченных клиентов (алерт менеджеру если >60 мин)
    app.job_queue.run_repeating(
        _check_unanswered_clients,
        interval=15 * 60,
        first=15 * 60,
    )
    # Каждый час: персональные касания клиентов (day1/day3/week1/week4/monday)
    app.job_queue.run_repeating(
        touchpoints.run_touchpoints_job,
        interval=60 * 60,
        first=60 * 60,
    )
    logger.info("Бот Валли запущен")
    app.run_polling()


if __name__ == "__main__":
    main()
