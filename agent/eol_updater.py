#!/usr/bin/env python3
"""
EOL updater for price.xlsx.

Reads articles from column C (rows 3–102), searches Cisco website for
End-of-Sale dates, writes found dates to column B.

Strategy:
  1. Try Cisco search page (fast, but JS-heavy — often yields nothing)
  2. Use DuckDuckGo HTML to locate a Cisco EoL bulletin URL
  3. Scrape the bulletin page and extract "End-of-Sale Date"

Skips rows where column B is already filled.
Saves file after every 10 processed rows and at the end.

Usage:
    python3 eol_updater.py            # process all rows 3–102
    python3 eol_updater.py --limit 5  # process first 5 non-empty articles
"""

import argparse
import logging
import os
import re
import time
import warnings
from datetime import datetime
from urllib.parse import parse_qs, quote, unquote, urlparse

import openpyxl
import requests
from bs4 import BeautifulSoup

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

PRICE_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "price.xlsx")
ROW_START = 3
ROW_END = 102
COL_ARTICLE = 3   # column C
COL_EOL = 2       # column B
PAUSE = 2         # seconds between HTTP requests
SAVE_EVERY = 10   # save file every N processed rows

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
)
logger = logging.getLogger(__name__)

_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
}

# ---------------------------------------------------------------------------
# Date parsing
# ---------------------------------------------------------------------------

_DATE_FORMATS = [
    "%B %d, %Y",    # January 31, 2024
    "%B %d %Y",     # January 31 2024
    "%b %d, %Y",    # Jan 31, 2024
    "%b. %d, %Y",   # Jan. 31, 2024
    "%d %B %Y",     # 31 January 2024
    "%d %b %Y",     # 31 Jan 2024
    "%m/%d/%Y",     # 01/31/2024
    "%d.%m.%Y",     # 31.01.2024
    "%Y-%m-%d",     # 2024-01-31
]

_EOS_PATTERNS = [
    r"End[- ]of[- ]Sale\s+Date\s*[:\-]\s*([A-Za-z]+\.?\s+\d{1,2},?\s+\d{4})",
    r"End[- ]of[- ]Sale\s*Date[:\-]\s*([A-Za-z]+\.?\s+\d{1,2},?\s+\d{4})",
    r"End[- ]of[- ]Sale[:\-]\s*([A-Za-z]+\.?\s+\d{1,2},?\s+\d{4})",
    r"EoS\s+Date\s*[:\-]\s*([A-Za-z]+\.?\s+\d{1,2},?\s+\d{4})",
    r"End[- ]of[- ]Life\s+Announcement\s*[:\-]\s*([A-Za-z]+\.?\s+\d{1,2},?\s+\d{4})",
    # Numeric dates
    r"End[- ]of[- ]Sale\s+Date\s*[:\-]\s*(\d{1,2}[./]\d{1,2}[./]\d{4})",
    r"End[- ]of[- ]Sale\s*Date[:\-]\s*(\d{1,2}[./]\d{1,2}[./]\d{4})",
]


def _parse_date(s: str) -> datetime | None:
    s = re.sub(r"\s+", " ", s.strip().rstrip(".").strip())
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def _extract_eos_date(html: str) -> datetime | None:
    """Extract the earliest End-of-Sale date from an HTML page."""
    text = BeautifulSoup(html, "html.parser").get_text(" ", strip=True)
    found: list[datetime] = []
    for pat in _EOS_PATTERNS:
        for m in re.finditer(pat, text, re.IGNORECASE):
            d = _parse_date(m.group(1))
            if d:
                found.append(d)
    return min(found) if found else None

# ---------------------------------------------------------------------------
# Article cleaning
# ---------------------------------------------------------------------------

_JUNK_SUFFIX_RE = re.compile(r"\s+(tech|used|ref|б\/у|new|like\s*new)\s*$", re.IGNORECASE)


def _clean_article(raw: str) -> str:
    """Strip whitespace, non-ASCII, and trailing condition labels."""
    s = raw.strip()
    # Remove non-breaking spaces and other non-ASCII control chars
    s = re.sub(r"[\xa0​­]+", "", s).strip()
    # Remove trailing condition qualifiers before searching
    s = _JUNK_SUFFIX_RE.sub("", s).strip()
    return s

# ---------------------------------------------------------------------------
# HTTP helpers
# ---------------------------------------------------------------------------

def _resolve_ddg_href(href: str) -> str:
    """Unwrap DuckDuckGo redirect link to the real URL."""
    if "duckduckgo.com/l/" in href:
        parsed = urlparse("https:" + href if href.startswith("//") else href)
        params = parse_qs(parsed.query)
        if "uddg" in params:
            return unquote(params["uddg"][0])
    return href


def _ddg_find_cisco_eol(article: str, session: requests.Session) -> str | None:
    """
    Search DuckDuckGo HTML for a Cisco EoL bulletin URL or date snippet.
    Returns either:
      - a full https:// URL to a Cisco EoL bulletin page
      - "__date__YYYY-MM-DD" if a date was found directly in the snippet
      - None if nothing found
    """
    q = f"site:cisco.com {article} end-of-sale eos-eol-notice"
    url = f"https://html.duckduckgo.com/html/?q={quote(q)}"
    try:
        r = session.get(url, timeout=15)
        if r.status_code != 200:
            return None
        soup = BeautifulSoup(r.text, "html.parser")

        # Try to extract date directly from result snippets (fast path)
        for snippet in soup.select(".result__snippet, .result__body, .result__a"):
            d = _extract_eos_date(str(snippet))
            if d:
                return f"__date__{d.strftime('%Y-%m-%d')}"

        # Look for Cisco EoL bulletin links
        for a_tag in soup.find_all("a", href=True):
            real = _resolve_ddg_href(a_tag["href"])
            if (
                "cisco.com" in real
                and "eos-eol" in real.lower()
                and real.startswith("http")
            ):
                return real
    except Exception as e:
        logger.warning("  DDG error for %s: %s", article, e)
    return None

# ---------------------------------------------------------------------------
# Main lookup
# ---------------------------------------------------------------------------

def find_eol_date(article: str, session: requests.Session) -> datetime | None:
    """Return End-of-Sale date for the article, or None if active / not found."""

    # Step 1: Try Cisco search directly (SPA — usually returns empty shell,
    #         but worth a quick attempt in case of partial server-side render)
    try:
        cisco_url = (
            f"https://search.cisco.com/search"
            f"?query={quote(article)}+end+of+sale&locale=enUS&cat=All"
        )
        r = session.get(cisco_url, timeout=12)
        if r.status_code == 200:
            d = _extract_eos_date(r.text)
            if d:
                logger.info("  → date from Cisco search: %s", d.strftime("%d.%m.%Y"))
                return d
    except Exception:
        pass
    time.sleep(PAUSE)

    # Step 2: DuckDuckGo → find Cisco EoL bulletin URL
    result = _ddg_find_cisco_eol(article, session)
    time.sleep(PAUSE)

    if not result:
        return None

    # Fast path: date was in the DDG snippet
    if result.startswith("__date__"):
        d = _parse_date(result[8:])
        if d:
            logger.info("  → date from DDG snippet: %s", d.strftime("%d.%m.%Y"))
        return d

    # Step 3: Scrape the Cisco EoL bulletin page
    logger.info("  → bulletin found: %s", result)
    try:
        r = session.get(result, timeout=15, allow_redirects=True)
        if r.status_code == 200:
            d = _extract_eos_date(r.text)
            if d:
                logger.info("  → date from bulletin page: %s", d.strftime("%d.%m.%Y"))
                return d
    except Exception as e:
        logger.warning("  Failed to scrape bulletin %s: %s", result, e)
    time.sleep(PAUSE)

    return None

# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Update EoL dates in price.xlsx")
    parser.add_argument("--limit", type=int, default=0,
                        help="Stop after N processed (non-skipped) articles (0 = all)")
    args = parser.parse_args()

    logger.info("Loading %s", PRICE_PATH)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(PRICE_PATH)
    ws = wb["Глав"]

    session = requests.Session()
    session.headers.update(_HEADERS)

    found_n = skipped_n = not_found_n = processed_n = 0

    for row in range(ROW_START, ROW_END + 1):
        raw_article = ws.cell(row=row, column=COL_ARTICLE).value
        if not raw_article or str(raw_article).strip() == "":
            continue

        article = _clean_article(str(raw_article))
        if not article:
            continue

        eol_cell = ws.cell(row=row, column=COL_EOL)
        existing = eol_cell.value
        if existing is not None and str(existing).strip() not in ("", "None", "nan"):
            logger.info("Row %3d: %-42s → SKIP (already: %s)", row, article, existing)
            skipped_n += 1
            continue

        logger.info("Row %3d: %-42s → searching...", row, article)
        eol_date = find_eol_date(article, session)
        processed_n += 1

        if eol_date:
            eol_cell.value = eol_date
            logger.info("Row %3d: %-42s → ✓ EoS: %s", row, article, eol_date.strftime("%d.%m.%Y"))
            found_n += 1
        else:
            logger.info("Row %3d: %-42s → active / not found", row, article)
            not_found_n += 1

        # Periodic save
        if processed_n % SAVE_EVERY == 0:
            wb.save(PRICE_PATH)
            logger.info("  [autosave after %d rows]", processed_n)

        if args.limit and processed_n >= args.limit:
            logger.info("Reached --limit %d, stopping.", args.limit)
            break

    wb.save(PRICE_PATH)
    wb.close()
    logger.info(
        "Done!  Found: %d | Active/not found: %d | Skipped: %d | Total processed: %d",
        found_n, not_found_n, skipped_n, processed_n,
    )


if __name__ == "__main__":
    main()
